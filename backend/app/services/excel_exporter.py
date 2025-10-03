"""
Excel file exporter for employee salary data.
Generates Excel files using a template to preserve formatting.
"""
import io
from typing import List
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from ..models.schemas import SelectEmployee


def export_employees_to_excel(employees: List[SelectEmployee]) -> bytes:
    """
    Export employees to Excel format using the template file and preserving its format.
    
    Args:
        employees: List of employee objects to export
        
    Returns:
        Bytes content of the Excel file
    """
    # Path to the template file - relative to this file's location
    # This file is at: backend/app/services/excel_exporter.py
    # Template is at: backend/templates/salary_template.xlsx
    current_file_dir = Path(__file__).parent  # app/services/
    backend_dir = current_file_dir.parent.parent  # backend/
    template_path = backend_dir / "templates" / "salary_template.xlsx"
    
    # Load the template Excel file
    workbook = load_workbook(template_path)
    worksheet: Worksheet = workbook.active
    
    # Start writing data from row 2 (row 1 has headers)
    start_row = 2
    
    # Save formatting from row 2 if it exists (use as template for all data rows)
    row_format_template = {}
    if worksheet.max_row >= 2:
        for col in range(1, 27):
            cell = worksheet.cell(row=2, column=col)
            row_format_template[col] = {
                'number_format': cell.number_format,
                'font': cell.font.copy() if cell.font else None,
                'alignment': cell.alignment.copy() if cell.alignment else None,
                'border': cell.border.copy() if cell.border else None,
                'fill': cell.fill.copy() if cell.fill else None
            }
    
    # Clear all existing data rows (keep header)
    # This ensures we don't have leftover data from the template
    if worksheet.max_row > 1:
        for row_to_clear in range(2, worksheet.max_row + 1):
            for col in range(1, 27):
                worksheet.cell(row=row_to_clear, column=col).value = None
    
    # Now write the new data
    for idx, emp in enumerate(employees, 1):
        row = start_row + idx - 1
        
        # Calculate He so: OT15*0.5 + OT20 + OT30*2
        he_so = (emp.ot15 or 0) * 0.5 + (emp.ot20 or 0) + (emp.ot30 or 0) * 2
        
        # Calculate total OT hours (must include he_so as per backend calculator)
        # Formula: ot15 + ot20 + ot30 + he_so
        total_ot_hours = (emp.ot15 or 0) + (emp.ot20 or 0) + (emp.ot30 or 0) + he_so
        
        # Calculate thu nhập ko tính thuế (PIT) - non-taxable income
        # Formula: personalRelief + dependentRelief + employeeInsurance
        pit_non_taxable_income = emp.personalRelief + emp.dependentRelief + emp.employeeInsurance
        
        # Fill in the data - just values, no formatting changes
        worksheet.cell(row=row, column=1).value = idx  # STT
        worksheet.cell(row=row, column=2).value = emp.employeeNo  # No.
        worksheet.cell(row=row, column=3).value = emp.name  # Name
        worksheet.cell(row=row, column=4).value = emp.salary  # Salary
        worksheet.cell(row=row, column=5).value = emp.augSalary  # Aug's salary
        worksheet.cell(row=row, column=6).value = emp.bonus  # Bonus
        worksheet.cell(row=row, column=7).value = emp.allowanceTax  # Allowance tính thuế
        worksheet.cell(row=row, column=8).value = emp.overtimePayPIT  # Over time Pay PIT
        worksheet.cell(row=row, column=9).value = emp.totalSalary  # Total Salary
        worksheet.cell(row=row, column=10).value = emp.dependants  # Dependants
        worksheet.cell(row=row, column=11).value = emp.personalRelief  # Personal relief
        worksheet.cell(row=row, column=12).value = emp.dependentRelief  # Dependent relief
        worksheet.cell(row=row, column=13).value = emp.assessableIncome  # Assessable income
        worksheet.cell(row=row, column=14).value = pit_non_taxable_income  # thu nhập ko tính thuế (PIT)
        worksheet.cell(row=row, column=15).value = emp.personalIncomeTax  # Personal Income tax
        worksheet.cell(row=row, column=16).value = emp.companyInsurance  # Insurance - Company's pay
        worksheet.cell(row=row, column=17).value = emp.employeeInsurance  # Insurance - Employee's pay
        worksheet.cell(row=row, column=18).value = emp.unionFee  # Đoàn phí
        worksheet.cell(row=row, column=19).value = emp.overtimePayNonPIT  # Over time none pay PIT
        worksheet.cell(row=row, column=20).value = emp.advance  # Trừ Adv
        worksheet.cell(row=row, column=21).value = emp.totalNetIncome  # Total Net Income
        worksheet.cell(row=row, column=22).value = he_so  # He so
        worksheet.cell(row=row, column=23).value = emp.ot15  # OT ( 1.5 % )
        worksheet.cell(row=row, column=24).value = emp.ot20  # OT( 2.0%)
        worksheet.cell(row=row, column=25).value = emp.ot30  # OT( 3.0%)
        worksheet.cell(row=row, column=26).value = total_ot_hours  # Total OT hours
        
        # Apply the saved formatting from template row 2 to all data rows
        if row_format_template:
            for col in range(1, 27):
                cell = worksheet.cell(row=row, column=col)
                fmt = row_format_template[col]
                if fmt['number_format']:
                    cell.number_format = fmt['number_format']
                if fmt['font']:
                    cell.font = fmt['font']
                if fmt['alignment']:
                    cell.alignment = fmt['alignment']
                if fmt['border']:
                    cell.border = fmt['border']
                if fmt['fill']:
                    cell.fill = fmt['fill']
    
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()